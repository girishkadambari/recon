"""
Unit tests: XLSXExportBuilder — pure openpyxl output, no DB, no S3.
"""
import io
import pytest
import openpyxl

from app.domain.services.xlsx_export_builder import build_xlsx, _fmt_decimal
from decimal import Decimal

SAMPLE_MATCHES = [
    {
        "id": "m-001", "match_strategy": "EXACT_ID", "confidence_score": 100,
        "status": "MATCHED", "source_record_id": "src-001", "source_table": "payment_records",
        "target_record_id": "tgt-001", "target_table": "bank_records",
        "amount_delta": Decimal("0"), "date_delta_days": 0, "review_note": None,
    },
    {
        "id": "m-002", "match_strategy": "AMOUNT_DATE", "confidence_score": 92,
        "status": "PENDING_REVIEW", "source_record_id": "src-002", "source_table": "payment_records",
        "target_record_id": "tgt-002", "target_table": "bank_records",
        "amount_delta": Decimal("0"), "date_delta_days": 2, "review_note": None,
    },
]

SAMPLE_EXCEPTIONS = [
    {
        "id": "e-001", "file_role": "SOURCE", "reason": "UNMATCHED_SOURCE",
        "status": "OPEN", "amount": Decimal("3500.00"), "currency": "INR",
        "record_id": "src-003", "record_table": "payment_records",
        "ai_explanation": "This payment was not found in the bank statement.",
        "resolution_note": None,
    },
]


class TestXLSXBuilder:

    def _parse(self, xlsx_bytes: bytes) -> openpyxl.Workbook:
        return openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    def test_returns_bytes(self):
        result = build_xlsx(
            run_name="Test Run", run_status="COMPLETED",
            run_date="2024-01-31", match_rate_pct=80,
            total_source_rows=3, total_target_rows=3,
            matched_count=2, exception_count=1,
            matches=SAMPLE_MATCHES, exceptions=SAMPLE_EXCEPTIONS,
        )
        assert isinstance(result, bytes)
        assert len(result) > 100  # non-trivially sized

    def test_has_three_sheets(self):
        result = build_xlsx(
            run_name="Test", run_status="COMPLETED",
            run_date="2024-01-31", match_rate_pct=75,
            total_source_rows=2, total_target_rows=2,
            matched_count=1, exception_count=1,
            matches=SAMPLE_MATCHES, exceptions=SAMPLE_EXCEPTIONS,
        )
        wb = self._parse(result)
        assert wb.sheetnames == ["Summary", "Matches", "Exceptions"]

    def test_summary_contains_run_name(self):
        result = build_xlsx(
            run_name="January Stripe Recon", run_status="COMPLETED",
            run_date="2024-01-31", match_rate_pct=90,
            total_source_rows=10, total_target_rows=10,
            matched_count=9, exception_count=1,
            matches=[], exceptions=[],
        )
        wb = self._parse(result)
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 15)]
        assert "January Stripe Recon" in values

    def test_matches_sheet_has_correct_rows(self):
        result = build_xlsx(
            run_name="Test", run_status="COMPLETED",
            run_date="2024-01-31", match_rate_pct=100,
            total_source_rows=2, total_target_rows=2,
            matched_count=2, exception_count=0,
            matches=SAMPLE_MATCHES, exceptions=[],
        )
        wb = self._parse(result)
        ws = wb["Matches"]
        assert ws.max_row == 3  # header + 2 data rows

    def test_exceptions_sheet_has_correct_rows(self):
        result = build_xlsx(
            run_name="Test", run_status="COMPLETED",
            run_date="2024-01-31", match_rate_pct=50,
            total_source_rows=2, total_target_rows=2,
            matched_count=1, exception_count=1,
            matches=[], exceptions=SAMPLE_EXCEPTIONS,
        )
        wb = self._parse(result)
        ws = wb["Exceptions"]
        assert ws.max_row == 2  # header + 1 data row

    def test_ai_explanation_in_exceptions(self):
        result = build_xlsx(
            run_name="Test", run_status="COMPLETED",
            run_date="2024-01-31", match_rate_pct=0,
            total_source_rows=1, total_target_rows=0,
            matched_count=0, exception_count=1,
            matches=[], exceptions=SAMPLE_EXCEPTIONS,
        )
        wb = self._parse(result)
        ws = wb["Exceptions"]
        # AI explanation is column 9
        assert "not found in the bank statement" in (ws.cell(row=2, column=9).value or "")

    def test_empty_matches_and_exceptions(self):
        result = build_xlsx(
            run_name="Empty Run", run_status="COMPLETED",
            run_date="2024-01-31", match_rate_pct=0,
            total_source_rows=0, total_target_rows=0,
            matched_count=0, exception_count=0,
            matches=[], exceptions=[],
        )
        wb = self._parse(result)
        assert wb["Matches"].max_row == 1   # headers only
        assert wb["Exceptions"].max_row == 1


class TestFmtDecimal:
    def test_decimal_to_float(self):
        assert _fmt_decimal(Decimal("5000.50")) == pytest.approx(5000.50)

    def test_none_returns_empty_string(self):
        assert _fmt_decimal(None) == ""

    def test_string_decimal(self):
        assert _fmt_decimal("1234.56") == pytest.approx(1234.56)
